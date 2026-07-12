"""Parse the Games Master List workbook into normalized, browser-ready JSON.

The workbook has three sheets with quirky encodings we normalize here:
  * dates are Excel serials (e.g. 46080)         -> ISO "YYYY-MM-DD"
  * ratings are 0..1 fractions (0.9 == 90%)      -> float 0..1 (frontend shows %)
  * several columns are 0/1 flags                -> real booleans
  * time columns are decimal hours (9.955 ~ 10h) -> float hours (frontend "Xh Ym")

We also STRIP the "Games On Order" sheet's private columns (home address,
order #, tracking #) so they never reach the unauthenticated frontend.

Output shape (consumed by static/app.js):

    {
      "games":     {"columns": [<col>...], "rows": [{key: value}...]},
      "completed": {...},
      "onOrder":   {...},
    }

where <col> = {key, label, type, facet, search, sort, primary}. Rows only carry
non-empty values (keeps the ~14.7k-row payload small under gzip).
"""

from __future__ import annotations

import datetime as _dt
import re

from openpyxl import load_workbook

import notes as notes_mod

# Excel's day 0. Serials count days after this; using 1899-12-30 (rather than
# 1900-01-01) absorbs Excel's phantom 1900-02-29 for every modern date.
_EXCEL_EPOCH = _dt.datetime(1899, 12, 30)


def _slug(header: str) -> str:
    """'Release Year' -> 'releaseYear', 'GameFAQs User Rating' -> 'gamefaqsUserRating'."""
    parts = re.split(r"[^0-9a-zA-Z]+", header.strip())
    parts = [p for p in parts if p]
    if not parts:
        return "col"
    first = parts[0].lower()
    return first + "".join(p[:1].upper() + p[1:] for p in parts[1:])


# ---- column schema -------------------------------------------------------
# Each entry: (header, type, facet, search, primary)
#   type:    text|date|rating|bool|hours|number|int|year
#   facet:   offer as a filter facet in the sidebar
#   search:  include in the free-text search index
#   primary: show as a column in the main table (else detail-drawer only)
# Unknown headers not listed here still pass through as plain text (so a new
# spreadsheet column never silently disappears) — except EXCLUDE headers.

_GAMES = [
    ("Title",                "text",   False, True,  True),
    ("Platform",             "text",   True,  False, True),
    ("Release Date",         "date",   False, False, True),
    ("Release Year",         "year",   True,  False, False),
    ("Release Region",       "text",   True,  False, False),
    ("Genre",                "text",   True,  True,  True),
    ("Franchise",            "text",   True,  True,  False),
    ("Publisher",            "text",   True,  True,  False),
    ("Developer",            "text",   True,  True,  False),
    ("Rating",               "rating", False, False, True),
    ("Metacritic Rating",    "rating", False, False, False),
    ("GameFAQs User Rating", "rating", False, False, False),
    ("Playing Status",       "text",   True,  False, True),
    ("Playing Progress",     "text",   False, False, False),
    ("Priority",             "text",   True,  False, False),
    ("Format",               "text",   True,  False, False),
    ("Owned",                "bool",   True,  False, True),
    ("Completed",            "bool",   True,  False, True),
    ("Wishlisted",           "bool",   True,  False, False),
    # Tri-state, not a bool (see _VALUE_LABELS): 1 playable, 0 unknown, -1 not.
    ("Playable",             "text",   True,  False, False),
    ("VR",                   "bool",   True,  False, False),
    ("DLC",                  "bool",   True,  False, False),
    # Tri-state translation status, not a bool (see _VALUE_LABELS).
    ("English",              "text",   True,  False, False),
    ("Condition",            "text",   False, False, False),
    ("Date Purchased",       "date",   False, False, False),
    ("Purchase Price",       "money",  False, False, False),
    ("Date Started",         "date",   False, False, False),
    ("Date Completed",       "date",   False, False, False),
    ("Completion Time",      "hours",  False, False, False),
    ("Estimated Time",       "hours",  False, False, False),
    ("Date Added",           "date",   False, False, False),
    ("File Size",            "text",   False, False, False),
    ("MAME Romset",          "text",   False, False, False),
    ("Notes",                "text",   False, True,  False),
    # Derived from Notes (see notes.py). The sheet packs eight different facts
    # into one free-text cell; these unpack it into things you can filter on.
    ("Digital Platform",     "text",   True,  False, False),
    ("Subscription",         "text",   True,  False, False),
    ("Limited Print",        "text",   True,  False, False),
    ("Edition",              "text",   True,  False, False),
    ("Required Accessory",   "text",   True,  False, False),
    ("Physical Media",       "text",   True,  False, False),
    ("Delisted",             "bool",   True,  False, False),
    ("Damaged",              "bool",   True,  False, False),
]

_COMPLETED = [
    ("Game",           "text",   False, True,  True),
    ("Platform",       "text",   True,  False, True),
    ("Release",        "date",   False, False, False),
    ("Region",         "text",   True,  False, False),
    ("Genre",          "text",   True,  True,  True),
    ("Franchise",      "text",   True,  True,  False),
    ("Publisher",      "text",   True,  True,  False),
    ("Developer",      "text",   True,  True,  False),
    ("Rating",         "rating", False, False, True),
    ("Critic Score",   "rating", False, False, False),
    ("Date",           "date",   False, False, True),   # date completed
    ("Started",        "date",   False, False, False),
    ("Play Time",      "hours",  False, False, True),
    ("Steam Deck",     "bool",   True,  False, False),
    ("Emulated",       "bool",   True,  False, False),
    ("VR",             "bool",   True,  False, False),
    ("Collection",     "text",   True,  False, False),
    ("#",              "int",    False, False, False),
    ("Notes",          "text",   False, True,  False),  # long-form review
]

_ON_ORDER = [
    ("Title",             "text",   False, True,  True),
    ("Platform",          "text",   True,  False, True),
    ("Vendor",            "text",   True,  True,  True),
    ("Status",            "text",   True,  False, True),
    ("Format",            "text",   True,  False, True),
    ("Price",             "money",  False, False, True),
    ("Ordered Date",      "date",   False, False, True),
    ("Estimated Release", "date",   False, False, True),
]

# Never emit these — private info that must not reach an unauthenticated page.
_EXCLUDE = {"Order #", "Address on Order", "Tracking #"}

# Map coded cell values to human labels (applies to column, facet, and drawer).
# Keyed by the slugged column key.
_REGIONS = {
    "NA": "North America", "EU": "Europe", "JP": "Japan", "AS": "Asia",
    "BR": "Brazil", "DE": "Germany", "FR": "France", "KO": "Korea",
    "SP": "Spain", "TW": "Taiwan", "IT": "Italy", "CN": "China",
    "SE": "Sweden", "AU": "Australia", "YU": "Yugoslavia",
}

# Coded columns → human labels. These mirror the enums in zdiemer/GamesMaster;
# several were previously coerced to bool, which silently collapsed states.
_VALUE_LABELS = {
    "playingStatus": {"1": "Playing", "0": "On Hold", "-1": "Up Next"},
    # Playability: UNKNOWN=0, PLAYABLE=1, UNPLAYABLE=-1. As a bool, bool(-1) was
    # True → unplayable games read "Yes" and unknown ones "No".
    "playable": {"1": "Yes", "0": "Unknown", "-1": "No"},
    # TranslationStatus: NONE=0, PARTIAL=1, COMPLETE=2. As a bool, Partial and
    # Full both collapsed to True. Blank = natively English (no entry).
    "english": {"0": "None", "1": "Partial", "2": "Full"},
    # ExcelRegion codes → full names.
    "releaseRegion": _REGIONS,
    "region": _REGIONS,
}


def _apply_label(key, value):
    mapping = _VALUE_LABELS.get(key)
    if not mapping:
        return value
    # Normalize numeric-ish cells ("1.0"/1.0/1 -> "1") before lookup.
    lookup = str(value)
    try:
        f = float(value)
        if f == int(f):
            lookup = str(int(f))
    except (TypeError, ValueError):
        pass
    return mapping.get(lookup, value)

# Match workbook sheets to our logical keys (by name, with positional fallback).
_SHEET_MAP = {
    "Games": ("games", _GAMES),
    "Games On Order": ("onOrder", _ON_ORDER),
    "Finished Games (Adtl Metadata)": ("completed", _COMPLETED),
}
_SHEET_ORDER = ["games", "onOrder", "completed"]
_SCHEMA_BY_KEY = {"games": _GAMES, "onOrder": _ON_ORDER, "completed": _COMPLETED}


def _to_number(value):
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        s = value.strip().replace(",", "")
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _serial_to_iso(value):
    """Excel serial (or an openpyxl datetime) -> 'YYYY-MM-DD'. Passthrough text."""
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.strftime("%Y-%m-%d")
    num = _to_number(value)
    if num is None:
        # Non-numeric date cell (e.g. "N/A", "TBD") — keep as-is.
        return str(value).strip() if value not in (None, "") else None
    if num <= 0:
        return None
    return (_EXCEL_EPOCH + _dt.timedelta(days=float(num))).strftime("%Y-%m-%d")


def _year_of(value):
    """Year from either a literal year cell (2026) or an Excel date serial."""
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.year
    num = _to_number(value)
    if num is None:
        return None
    n = int(num)
    if 1000 <= n <= 9999:
        return n  # already a plain 4-digit year
    iso = _serial_to_iso(num)  # otherwise treat as a date serial
    if iso and re.match(r"^\d{4}-", iso):
        return int(iso[:4])
    return None


def _coerce(value, ctype):
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None

    if ctype == "bool":
        num = _to_number(value)
        return bool(num) if num is not None else None
    if ctype == "rating":
        num = _to_number(value)
        return round(num, 4) if num is not None else None
    if ctype in ("hours", "number", "money"):
        num = _to_number(value)
        if num is None:
            return None
        return round(num, 4)
    if ctype == "int":
        num = _to_number(value)
        return int(num) if num is not None else None
    if ctype == "year":
        return _year_of(value)
    if ctype == "date":
        return _serial_to_iso(value)
    # text
    s = str(value).strip()
    return s or None


def _columns_for(schema):
    cols = []
    for header, ctype, facet, search, primary in schema:
        cols.append(
            {
                "key": _slug(header),
                "label": header,
                "type": ctype,
                "facet": facet,
                "search": search,
                "sort": True,
                "primary": primary,
            }
        )
    return cols


def _parse_sheet(ws, schema):
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        return {"columns": _columns_for(schema), "rows": []}

    headers = [str(h).strip() if h is not None else "" for h in headers]
    typemap = {h: (t, _slug(h)) for h, t, *_ in schema}

    out_rows = []
    for raw in rows_iter:
        record = {}
        has_title = False
        for idx, header in enumerate(headers):
            if not header or header in _EXCLUDE:
                continue
            value = raw[idx] if idx < len(raw) else None
            ctype, key = typemap.get(header, ("text", _slug(header)))
            coerced = _coerce(value, ctype)
            if coerced is None or coerced == "":
                continue
            if key in _VALUE_LABELS:
                coerced = _apply_label(key, coerced)
            record[key] = coerced
            if idx == 0:
                has_title = True
        # Skip fully blank / title-less rows (trailing spreadsheet padding).
        if record and has_title:
            out_rows.append(record)

    return {"columns": _columns_for(schema), "rows": out_rows}


def _inject_notes_fields(sheet: dict):
    """Unpack the Notes cell into the columns declared in the schema.

    Notes is one free-text field carrying eight different facts (storefront,
    subscription, boutique label, edition, delisted, damaged, accessory, media).
    Each row gets whichever of them its Notes value encodes — see notes.py, which
    ports GamesMaster's __process_notes, order and all.
    """
    for row in sheet["rows"]:
        for k, v in notes_mod.process(row.get("notes")).items():
            row[k] = v


def _inject_release_year(sheet: dict, src_key: str):
    """Add a facetable Release Year column derived from a date column (for the
    Completed sheet, which stores only a full release date, not a year)."""
    cols = sheet["columns"]
    if any(c["key"] == "releaseYear" for c in cols):
        return
    idx = next((i for i, c in enumerate(cols) if c["key"] == src_key), None)
    if idx is None:
        return
    cols.insert(
        idx + 1,
        {"key": "releaseYear", "label": "Release Year", "type": "year",
         "facet": True, "search": False, "sort": True, "primary": False},
    )
    for row in sheet["rows"]:
        v = row.get(src_key)
        if isinstance(v, str) and re.match(r"^\d{4}-", v):
            row["releaseYear"] = int(v[:4])


def parse_workbook(data: bytes) -> dict:
    """Parse xlsx bytes -> {games|onOrder|completed: {columns, rows}}."""
    import io

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    result = {}
    try:
        sheetnames = list(wb.sheetnames)
        for name in sheetnames:
            mapped = _SHEET_MAP.get(name)
            if mapped:
                key, schema = mapped
                result[key] = _parse_sheet(wb[name], schema)
        # Positional fallback for any logical sheet we didn't match by name.
        for i, key in enumerate(_SHEET_ORDER):
            if key not in result and i < len(sheetnames):
                result[key] = _parse_sheet(wb[sheetnames[i]], _SCHEMA_BY_KEY[key])
    finally:
        wb.close()

    # Guarantee all three keys exist even if a sheet was missing.
    for key in _SHEET_ORDER:
        result.setdefault(key, {"columns": _columns_for(_SCHEMA_BY_KEY[key]), "rows": []})

    # Notes is one cell doing eight jobs; unpack it into filterable columns.
    _inject_notes_fields(result["games"])

    # Completed sheet only carries a full release date — derive a year facet.
    _inject_release_year(result["completed"], "release")

    # Early Access / TBD games have no numeric year; surface the release-date
    # label (e.g. "Early Access") as the Release Year facet value so they're
    # filterable and sortable rather than dropping out of the year facet.
    for r in result["games"]["rows"]:
        if not r.get("releaseYear"):
            rd = r.get("releaseDate")
            if isinstance(rd, str) and not re.match(r"^\d{4}-", rd):
                r["releaseYear"] = rd
    return result
